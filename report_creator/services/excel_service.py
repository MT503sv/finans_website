from io import BytesIO
import pandas as pd
from openpyxl.utils import get_column_letter


def generate_excel(data, insights=None):

    buffer = BytesIO()

    if not data:
        data = [{"message": "No data available"}]

    df = pd.DataFrame(data)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        df.to_excel(writer, index=False, sheet_name="Report")

        ws = writer.sheets["Report"]

        if ws.max_column > 0:

            for col in ws.columns:

                max_len = 0

                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))

                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max_len + 3

    buffer.seek(0)
    return buffer