import os
import re
import asyncio
import uuid
import base64
from flask import Flask, request, jsonify
from flask_cors import CORS
from groq import Groq
from prisma import Prisma
from dotenv import load_dotenv

load_dotenv(override=True)

os.environ.setdefault("PRISMA_SCHEMA_PATH", "prisma/schema.python.prisma")

app = Flask(__name__)

CORS(app, resources={
    r"/*": {
        "origins": [
            "http://localhost:3000",
            "http://127.0.0.1:3000",
            "https://finans-website.onrender.com",
            "https://finanswebsite-production-7320.up.railway.app"
        ]
    }
})

@app.route("/", methods=["GET"])
def home():
    return jsonify({
        "status": "online",
        "message": "Kuali API running successfully"
    })

GROQ_API_KEY = os.getenv("GROQ_API_KEY")
if not GROQ_API_KEY:
    raise RuntimeError("Missing GROQ_API_KEY environment variable in .env file")

client = Groq(api_key=GROQ_API_KEY)

CALC_PATTERN = re.compile(
    r'(\d+)\s*[+-]\s*(\d+)\s*=\s*(\d+)|(\d+)\s*[+-]\s*(\d+)\s*=\s*\$(\d+)',
    re.IGNORECASE
)
LATEX_PATTERN = re.compile(
    r'\\\[\s*\\text\{.+\}\s*=\s*(\d+\s*-\s*\d+\s*=\s*\d+)\s*\]'
)

# ─────────────────────────────────────────────
# DB CONTEXT — filtered by user_id (Clerk ID)
# ─────────────────────────────────────────────

async def get_database_context(user_id: str):
    prisma = Prisma()
    await prisma.connect()
    try:
        from datetime import datetime, timezone, timedelta

        # Use El Salvador time (UTC-6) only to get the correct local date string.
        # But query the DB in pure UTC (midnight to midnight) because sales are
        # stored at 00:00:00 UTC (old records) or 12:00:00 UTC (new records) —
        # both fall within 00:00 UTC → 23:59 UTC of the same calendar day.
        sv_tz     = timezone(timedelta(hours=-6))
        now_sv    = datetime.now(sv_tz)
        today_str = now_sv.strftime('%Y-%m-%d')

        # Build the UTC window using the local date so it matches the user's day
        today_utc_start = datetime(now_sv.year, now_sv.month, now_sv.day,
                                   0, 0, 0, 0, tzinfo=timezone.utc)
        today_utc_end   = datetime(now_sv.year, now_sv.month, now_sv.day,
                                   23, 59, 59, 999999, tzinfo=timezone.utc)
        today_start = today_utc_start
        today_end   = today_utc_end

        # Verify the user exists
        user = await prisma.user.find_unique(where={"id": user_id})

        if not user:
            return f"=== KUALI FINANCIAL APPLICATION DATA ===\nToday's date: {today_str}\n\nNo data found for this user.\n"

        # ── ALL-TIME totals ──────────────────────────────────────────────
        all_sales    = await prisma.sales.find_many(where={"user_id": user_id})
        all_incomes  = await prisma.incomes.find_many(where={"user_id": user_id})
        all_outflows = await prisma.outflows.find_many(where={"user_id": user_id})

        total_sales_alltime    = sum(float(s.price_of_item * s.quantity_of_sold_items) for s in all_sales)
        total_incomes_alltime  = sum(float(i.amount) for i in all_incomes)
        total_outflows_alltime = sum(float(o.amount) for o in all_outflows)
        count_sales_alltime    = sum(s.quantity_of_sold_items for s in all_sales)

        # ── TODAY totals ─────────────────────────────────────────────────
        today_sales    = await prisma.sales.find_many(
            where={"user_id": user_id, "date": {"gte": today_start, "lte": today_end}},
            order={'date': 'desc'}
        )
        today_incomes  = await prisma.incomes.find_many(
            where={"user_id": user_id, "date": {"gte": today_start, "lte": today_end}},
            order={'date': 'desc'}
        )
        today_outflows = await prisma.outflows.find_many(
            where={"user_id": user_id, "date": {"gte": today_start, "lte": today_end}},
            order={'date': 'desc'}
        )

        total_sales_today    = sum(float(s.price_of_item * s.quantity_of_sold_items) for s in today_sales)
        total_incomes_today  = sum(float(i.amount) for i in today_incomes)
        total_outflows_today = sum(float(o.amount) for o in today_outflows)
        count_sales_today    = sum(s.quantity_of_sold_items for s in today_sales)
        count_transactions_today = len(today_sales)

        # ── Build context ────────────────────────────────────────────────
        contexto  = f"=== KUALI FINANCIAL APPLICATION DATA ===\n"
        contexto += f"Today's date: {today_str}\n\n"
        contexto += f"USER: {user.name} ({user.email})\n\n"

        contexto += "── ALL-TIME SUMMARY ──\n"
        contexto += f"Total items sold (all time): {count_sales_alltime}\n"
        contexto += f"Total sales revenue (all time): ${total_sales_alltime:.2f}\n"
        contexto += f"Total other income (all time): ${total_incomes_alltime:.2f}\n"
        contexto += f"Total expenses (all time): ${total_outflows_alltime:.2f}\n\n"

        contexto += f"── TODAY ({today_str}) SUMMARY ──\n"
        contexto += f"Sale transactions today: {count_transactions_today}\n"
        contexto += f"Total items sold today: {count_sales_today}\n"
        contexto += f"Total sales revenue today: ${total_sales_today:.2f}\n"
        contexto += f"Total other income today: ${total_incomes_today:.2f}\n"
        contexto += f"Total expenses today: ${total_outflows_today:.2f}\n"
        contexto += f"Net profit today: ${(total_sales_today + total_incomes_today - total_outflows_today):.2f}\n\n"

        contexto += f"── TODAY SALES DETAIL ({today_str}) ──\n"
        if today_sales:
            for s in today_sales:
                fecha_str  = s.date.strftime('%Y-%m-%d %H:%M') if s.date else 'N/A'
                total_item = s.quantity_of_sold_items * s.price_of_item
                contexto += (
                    f"- {s.quantity_of_sold_items}x '{s.item_sold}' "
                    f"@ ${float(s.price_of_item):.2f} each = ${total_item:.2f} | {fecha_str}\n"
                )
        else:
            contexto += "No sales recorded today.\n"

        contexto += f"\n── TODAY INCOMES DETAIL ({today_str}) ──\n"
        if today_incomes:
            for i in today_incomes:
                fecha_str = i.date.strftime('%Y-%m-%d %H:%M') if i.date else 'N/A'
                contexto += f"- {i.income_type} | {i.description or '---'} | ${float(i.amount):.2f} | {fecha_str}\n"
        else:
            contexto += "No other income recorded today.\n"

        contexto += f"\n── TODAY EXPENSES DETAIL ({today_str}) ──\n"
        if today_outflows:
            for o in today_outflows:
                fecha_str = o.date.strftime('%Y-%m-%d %H:%M') if o.date else 'N/A'
                contexto += f"- {o.outflow_type} | {o.description or '---'} | ${float(o.amount):.2f} | {fecha_str}\n"
        else:
            contexto += "No expenses recorded today.\n"

        contexto += "\n── RECENT SALES (last 50, all dates) ──\n"
        recent_sales = await prisma.sales.find_many(
            where={"user_id": user_id},
            order={'date': 'desc'},
            take=50
        )
        for s in recent_sales:
            fecha_str  = s.date.strftime('%Y-%m-%d') if s.date else 'N/A'
            total_item = s.quantity_of_sold_items * s.price_of_item
            contexto += (
                f"- {s.quantity_of_sold_items}x '{s.item_sold}' "
                f"= ${total_item:.2f} | {fecha_str}\n"
            )

        contexto += "\n── RECENT INCOMES (last 50) ──\n"
        recent_incomes = await prisma.incomes.find_many(
            where={"user_id": user_id},
            order={'date': 'desc'},
            take=50
        )
        for i in recent_incomes:
            fecha_str = i.date.strftime('%Y-%m-%d') if i.date else 'N/A'
            contexto += f"- {i.income_type} | {i.description or '---'} | ${float(i.amount):.2f} | {fecha_str}\n"

        contexto += "\n── RECENT EXPENSES (last 50) ──\n"
        recent_outflows = await prisma.outflows.find_many(
            where={"user_id": user_id},
            order={'date': 'desc'},
            take=50
        )
        for o in recent_outflows:
            fecha_str = o.date.strftime('%Y-%m-%d') if o.date else 'N/A'
            contexto += f"- {o.outflow_type} | {o.description or '---'} | ${float(o.amount):.2f} | {fecha_str}\n"

        return contexto
    finally:
        await prisma.disconnect()


def build_system_prompt(db_context: str) -> str:
    return f"""You are Kuali, an intelligent financial assistant.

Your specialty is:
- Analyzing purchases, sales, income, and expenses of the user.
- Giving advice on personal and business finances.
- Answering questions by analyzing existing transactions from the database.

RULES:
1. ALWAYS use the real data below when asked about finances, income, or sales.
2. If asked about something outside of finances, answer briefly but clarify that your area is business finances.
3. CRITICAL: Detect the language of the user's message and respond EXCLUSIVELY in that language. If the user writes in Spanish, respond in Spanish. If the user writes in English, respond in English. Never mix languages.
4. Be precise with numbers and monetary calculations.
5. CRITICAL — QUANTITY RULES (read carefully):
   - The context has a section "TODAY SUMMARY" with "Total items sold today: X". That X is the ONLY correct answer for "how many did I sell today".
   - The context has "ALL-TIME SUMMARY" with "Total items sold (all time): X". That X is the ONLY correct answer for all-time quantities.
   - NEVER count the number of lines in RECENT SALES or TODAY SALES DETAIL to answer quantity questions. Those sections show transactions, not individual item counts.
   - Example: if TODAY SUMMARY says "Total items sold today: 47", answer 47. Do NOT count the detail lines and answer 3 or 5.
   - Each detail line like "10x 'pupusas' @ $1.00 = $10.00" means 10 individual items. The number before "x" is the quantity, not 1.
   - If the user asks "how many pupusas did I sell", sum ONLY the lines where item_sold is 'pupusas' using the quantity before "x". Never count lines.
6. Only discuss data that belongs to this user. Never reference other users.

{db_context}"""


# ─────────────────────────────────────────────
# CHAT ENDPOINT
# ─────────────────────────────────────────────

@app.route('/chat', methods=['POST'])
def chat():
    data = request.get_json()
    user_question = data.get("question")
    user_id       = data.get("user_id")

    if not user_question:
        return jsonify({"error": "No question was provided"}), 400
    if not user_id:
        return jsonify({"error": "No user_id was provided"}), 400

    try:
        db_context    = asyncio.run(get_database_context(str(user_id)))
        system_prompt = build_system_prompt(db_context)

        chat_completion = client.chat.completions.create(
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user",   "content": user_question}
            ],
            model="llama-3.1-8b-instant",
            temperature=0.4
        )

        response_content = chat_completion.choices[0].message.content
        response_content = LATEX_PATTERN.sub(r'\1', response_content)

        def format_calc(match):
            if match.group(1):
                return f"## **${match.group(1)} - ${match.group(2)} = ${match.group(3)}**\n\n"
            return f"## **${match.group(4)} - ${match.group(5)} = ${match.group(6)}**\n\n"

        response_content = CALC_PATTERN.sub(format_calc, response_content)

        return jsonify({"response": response_content}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────
# CLIENTES ENDPOINT (kept for compatibility)
# ─────────────────────────────────────────────

@app.route('/clientes', methods=['GET'])
def get_clients():
    async def fetch():
        prisma = Prisma()
        await prisma.connect()
        try:
            users = await prisma.user.find_many(include={'sales': True})
            return [
                {
                    "name":           u.name,
                    "purchase_count": len(u.sales),
                    "total_spent":    sum(float(s.price_of_item * s.quantity_of_sold_items) for s in u.sales)
                }
                for u in users
            ]
        finally:
            await prisma.disconnect()

    try:
        result = asyncio.run(fetch())
        return jsonify(result), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────
# PDF HELPERS
# ─────────────────────────────────────────────

def pdf_header(c, report_type, start_date, end_date, width, height):
    from reportlab.lib import colors

    c.setFillColor(colors.HexColor("#010221"))
    c.rect(0, height - 80, width, 80, fill=1, stroke=0)

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(40, height - 45, f"Finans — {report_type.capitalize()} Report")

    c.setFont("Helvetica", 11)
    c.drawString(40, height - 65, f"Period: {start_date}  →  {end_date}")

    c.setFillColor(colors.black)

def pdf_table(c, data, headers, col_widths, y_start, width, height):
    from reportlab.lib import colors

    ROW_H    = 22
    HEADER_H = 26
    MARGIN   = 40
    x_start  = MARGIN

    c.setFillColor(colors.HexColor("#F4F5F7"))
    c.rect(x_start, y_start - HEADER_H, width - 2 * MARGIN, HEADER_H, fill=1, stroke=0)

    c.setFillColor(colors.HexColor("#010221"))
    c.setFont("Helvetica-Bold", 10)
    x = x_start + 6
    for i, h in enumerate(headers):
        c.drawString(x, y_start - 17, h)
        x += col_widths[i]

    y = y_start - HEADER_H

    c.setStrokeColor(colors.HexColor("#CCCCCC"))
    c.line(x_start, y, x_start + width - 2 * MARGIN, y)

    c.setFont("Helvetica", 9)
    for idx, row in enumerate(data):
        if y < 60:
            c.showPage()
            pdf_header(c, "", "", "", width, height)
            y = height - 100

        if idx % 2 == 0:
            c.setFillColor(colors.HexColor("#FAFAFA"))
            c.rect(x_start, y - ROW_H, width - 2 * MARGIN, ROW_H, fill=1, stroke=0)

        c.setFillColor(colors.black)
        x = x_start + 6
        for i, val in enumerate(row):
            c.drawString(x, y - 15, str(val))
            x += col_widths[i]

        c.setStrokeColor(colors.HexColor("#EEEEEE"))
        c.line(x_start, y - ROW_H, x_start + width - 2 * MARGIN, y - ROW_H)

        y -= ROW_H

    return y

def pdf_total(c, label, value, y, width):
    from reportlab.lib import colors

    MARGIN = 40
    c.setFillColor(colors.HexColor("#010221"))
    c.rect(MARGIN, y - 28, width - 2 * MARGIN, 28, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(MARGIN + 6, y - 18, label)
    c.drawRightString(width - MARGIN - 6, y - 18, value)
    c.setFillColor(colors.black)


# ─────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────

def excel_style(ws, headers, rows, col_widths):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DARK         = "010221"
    ALT          = "FAFAFA"
    BORDER_COLOR = "CCCCCC"

    thin   = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    ws.row_dimensions[1].height = 28

    for r_idx, row in enumerate(rows, 2):
        ws.append(row)
        fill_color = ALT if r_idx % 2 == 0 else "FFFFFF"
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[r_idx].height = 20

    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ─────────────────────────────────────────────
# GENERATE REPORT ENDPOINT — filtered by user_id
# ─────────────────────────────────────────────

@app.route('/generate_report', methods=['POST'])
def generate_report():

    async def fetch(user_id: str, report_type, start_date, end_date):
        prisma = Prisma()
        await prisma.connect()
        try:
            base_where = {
                "user_id": user_id,
                "date": {"gte": start_date, "lte": end_date}
            }

            if report_type == "incomes":
                rows = await prisma.incomes.find_many(where=base_where)
                data = [
                    {
                        "type":        r.income_type,
                        "description": r.description or "---",
                        "amount":      f"${float(r.amount):.2f}",
                        "date":        r.date.strftime('%Y-%m-%d') if r.date else "N/A",
                    }
                    for r in rows
                ]

            elif report_type == "expenses":
                rows = await prisma.outflows.find_many(where=base_where)
                data = [
                    {
                        "type":        r.outflow_type,
                        "description": r.description or "---",
                        "amount":      f"${float(r.amount):.2f}",
                        "date":        r.date.strftime('%Y-%m-%d') if r.date else "N/A",
                    }
                    for r in rows
                ]

            elif report_type == "sales":
                rows = await prisma.sales.find_many(where=base_where)
                data = [
                    {
                        "item":  r.item_sold,
                        "qty":   r.quantity_of_sold_items,
                        "price": f"${float(r.price_of_item):.2f}",
                        "total": f"${float(r.price_of_item * r.quantity_of_sold_items):.2f}",
                        "date":  r.date.strftime('%Y-%m-%d') if r.date else "N/A",
                    }
                    for r in rows
                ]

            elif report_type == "profit":
                ingresos = await prisma.incomes.find_many(where=base_where)
                egresos  = await prisma.outflows.find_many(where=base_where)
                ventas   = await prisma.sales.find_many(where=base_where)
                total_in  = sum(float(r.amount) for r in ingresos) + sum(float(r.price_of_item * r.quantity_of_sold_items) for r in ventas)
                total_out = sum(float(r.amount) for r in egresos)
                profit    = total_in - total_out
                data = [
                    {"concept": "Total Income",   "amount": f"${total_in:.2f}"},
                    {"concept": "Total Expenses", "amount": f"${total_out:.2f}"},
                    {"concept": "Net Profit",     "amount": f"${profit:.2f}"},
                ]
            else:
                data = []

            return data
        finally:
            await prisma.disconnect()

    try:
        body        = request.get_json()
        user_id     = body.get("user_id")
        report_type = body.get("report_type", "incomes")
        start_date  = body.get("start_date")
        end_date    = body.get("end_date")

        if not user_id:
            return jsonify({"error": "No user_id was provided"}), 400

        from datetime import datetime, timezone, timedelta

        # FIX: Query the DB using midnight-to-midnight UTC for the selected dates.
        # Sales are stored at 00:00:00 UTC (old records) or 12:00:00 UTC (new
        # records via Date.UTC noon fix) — both fall within the same UTC calendar
        # day, so this range captures everything correctly without any offset shift.
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(
            hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc
        )
        end_dt   = datetime.strptime(end_date, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc
        )

        print(f"DEBUG report: user={user_id} type={report_type} "
              f"start_sv={start_date} end_sv={end_date} "
              f"start_utc={start_dt} end_utc={end_dt}")

        data = asyncio.run(fetch(str(user_id), report_type, start_dt, end_dt))

        print(f"DEBUG report: {len(data)} rows returned")

        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas as pdf_canvas
        import openpyxl

        report_id  = str(uuid.uuid4())[:8]
        pdf_path   = f"static/{report_id}.pdf"
        excel_path = f"static/{report_id}.xlsx"
        os.makedirs("static", exist_ok=True)

        width, height = letter

        config = {
            "incomes":  {
                "headers":    ["Type", "Description", "Amount", "Date"],
                "keys":       ["type", "description", "amount", "date"],
                "col_widths": [100, 220, 90, 90],
                "excel_w":    [20, 40, 18, 18],
            },
            "expenses": {
                "headers":    ["Type", "Description", "Amount", "Date"],
                "keys":       ["type", "description", "amount", "date"],
                "col_widths": [100, 220, 90, 90],
                "excel_w":    [20, 40, 18, 18],
            },
            "sales": {
                "headers":    ["Item", "Qty", "Unit Price", "Total", "Date"],
                "keys":       ["item", "qty", "price", "total", "date"],
                "col_widths": [180, 50, 90, 90, 90],
                "excel_w":    [35, 10, 18, 18, 18],
            },
            "profit": {
                "headers":    ["Concept", "Amount"],
                "keys":       ["concept", "amount"],
                "col_widths": [350, 150],
                "excel_w":    [35, 20],
            },
        }

        cfg        = config.get(report_type, config["incomes"])
        headers    = cfg["headers"]
        keys       = cfg["keys"]
        col_widths = cfg["col_widths"]
        excel_w    = cfg["excel_w"]

        rows_for_table = [[row[k] for k in keys] for row in data]

        # PDF
        c = pdf_canvas.Canvas(pdf_path, pagesize=letter)
        pdf_header(c, report_type, start_date, end_date, width, height)

        y = height - 100
        y = pdf_table(c, rows_for_table, headers, col_widths, y, width, height)

        if report_type in ("incomes", "expenses") and data:
            total_val = sum(float(r["amount"].replace("$", "")) for r in data)
            pdf_total(c, "TOTAL", f"${total_val:.2f}", y - 10, width)

        c.save()

        # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.capitalize()
        excel_style(ws, headers, rows_for_table, excel_w)
        wb.save(excel_path)

        # Base64
        with open(pdf_path, "rb") as f:
            pdf_b64 = base64.b64encode(f.read()).decode("utf-8")
        with open(excel_path, "rb") as f:
            excel_b64 = base64.b64encode(f.read()).decode("utf-8")

        os.remove(pdf_path)
        os.remove(excel_path)

        return jsonify({
            "pdf":   f"data:application/pdf;base64,{pdf_b64}",
            "excel": f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_b64}",
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)